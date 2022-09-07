from unittest import case
from openpyxl import load_workbook

wb = load_workbook(r"E:\Excle_test\test.xlsx")

sh = wb["用户登录"]

# 只读取第一行
keys = []
for col_index in range(1,sh.max_column + 1) :
    keys.append(sh.cell(1, col_index).value)
print(keys)


for row_index in range(2, sh.max_row + 1) :
    values = []
    for col_index in range(1,sh.max_column + 1) :
        values.append(sh.cell(row_index,col_index).value)
    case = dict(zip(keys,values))
    print(case)

