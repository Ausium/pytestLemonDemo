"""
======================
Author: 柠檬班-小简
Time: 2021/4/23 20:23
Project: day5
Company: 湖南零檬信息技术有限公司
======================
"""
"""

1、安装
    pip install openpyxl

row  行
column 列
"""
from openpyxl import load_workbook

# excel的文件路径
excel_path = r"D:\Pychram-Workspace\py37-接口自动化\day5\testdatas\测试用例.xlsx"

# 1、加载一个excel，得到工作薄 Workbook
wb = load_workbook(excel_path)

# 2、选择一个表单- 通过表单名 Sheet
sh = wb["注册接口"]

# 3、在选择的表单当中，读取某个单元格的数据、修改/写入数据到某个单元格  Cell
# 行号和列号都是从1开始
# 读取
cell_value = sh.cell(2, 3).value  # 读取
print(cell_value)

# 得到当前sheet的总行号，总列号
row_nums = sh.max_row
col_nums = sh.max_column

# 只读取第一行(作为key)
# 行号是1  通过代码自动得到列号
# keys = []
# for col_index in range(1, sh.max_column + 1):
#     keys.append(sh.cell(1, col_index).value)
# print(keys)

# # 遍历行号，取第一行
# for row_index in range(2, sh.max_row + 1):
#     values = []
#     # 在每一行里面，从第1列开始，获取所有列的值
#     for col_index in range(1, sh.max_column + 1):
#         values.append(sh.cell(row_index, col_index).value)
#     # keys和values打包 - zip函数
#     case = dict(zip(keys, values))
#     print(case)

# 方式二 -
data = list(sh.values)
print(data)
keys = data[0] # 获取所有的列名
all_data = []
for row in data[1:]:
    row_dict = dict(zip(keys,row))
    all_data.append(row_dict)


# # 列表推导式  列表名 = [值 表达式]
# keys = [sh.cell(1, col_index).value for col_index in range(1, sh.max_column + 1)]


# # 读取所有行
# for row in sh.rows:
#     # print(row)
#     for item in row:
#         print(item.value, end="  ")
#     print()

# # 给某个单元格写入值
# sh.cell(2, 3).value = "get"
#
# # 一旦做了修改，就要保存
# # filename如果不是打开的excel文件，那就是另存为
# # 如果是打开的excel文件，保存到原文件中
# # 保存的时候，要保证没有其它程序在使用当前文件。否则会报Permission Error
# wb.save(excel_path)

