
from openpyxl import load_workbook



class Myexcle :
    def __init__(self,excle_path,sheet_name):
        # 1、加载一个excle 
        wb = load_workbook(excle_path)
        #  2、选择一个表单
        self.sh = wb[sheet_name]


    def read_data(self):
        #获取表头信息
        keys = []
        for col_index in range(1,self.sh.max_column + 1) :
            keys.append(self.sh.cell(1, col_index).value)
        # print(keys)

        #获取每条测试用例的数据
        all_values = []
        for row_index in range(2, self.sh.max_row + 1) :
            all_data = []
            for col_index in range(1,self.sh.max_column + 1) :
                all_data.append(self.sh.cell(row_index,col_index).value)

                #赋予的这个值，只会给一次，需要创建一个列表，把值放入这个列表中，才能得到全部的数值
                case = dict(zip(keys,all_data))            
            all_values.append(case)
        return all_values
        

if __name__ == "__main__" :
    #excle路径
    excle_path = r"E:\BaiduNetdiskDownload\InterfaceAuto\Excle_test\test.xlsx"
    me = Myexcle(excle_path,"用户登录")
    cases = me.read_data()
    for case in cases :
        print(case)


            

